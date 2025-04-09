import os
import pandas as pd
import numpy as np
from .short_utilities import *
import re
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import LabelEncoder
import calendar
import nltk
from nltk.stem import WordNetLemmatizer
from nltk.tokenize import word_tokenize
from sklearn.base import BaseEstimator, TransformerMixin
import xgboost as xgb
from sklearn.preprocessing import FunctionTransformer
from sklearn.pipeline import FeatureUnion
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import StandardScaler
from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
import imblearn

# from imblearn.under_sampling import RandomUnderSampler
# from imblearn.over_sampling import RandomOverSampler


def locate_train(min_store, excel_paths={}):
    # * train_excels: dict - {store_code: excel_filename} where excel_filenameis the latest file for each store
    # * excel_paths: dict - {excel_filename: excel_path} where excel_path is the path to the excel file

    excels, excel_paths = locate_excel("function/training_data/", excel_paths)

    # Group excels by store
    store_excels = defaultdict(list)
    for xl in excels:
        store = xl.split("-")[0]
        if store in min_store:
            store_excels[store].append(xl)

    train_excels = {}
    for store in min_store:
        if store in store_excels:
            # Sort excels for each store by yearmonth (descending order)
            sorted_excels = sorted(
                store_excels[store],
                key=lambda x: int(re.findall(r"\d{6}", x)[0]),
                reverse=True,
            )
            # Select the excel with the largest yearmonth
            train_excels[store] = sorted_excels[0]
        else:
            print(f"Warning: No excel file found for store {store}")
    excel_paths = {excel: excel_paths[excel] for excel in train_excels.values()}
    return train_excels, excel_paths


def locate_test(path, excel_path, delimiter="-", rawdata_keyword="Adv Data"):
    #! SHOULD BE REMOVE
    excels = {}
    stores = []
    # global excel_path

    for file in os.listdir(path):
        if not file.startswith("~$"):
            if file.endswith(".xlsx"):
                if re.search(rawdata_keyword, file):
                    names = file.split(delimiter)
                    stores.append(names[0])
                    excels[names[0]] = file
                    excel_path[file] = path + "/" + file
    stores = list(excels.keys())
    return excels, excel_path, stores, names


def merge_train(excel_paths, train_excels, data_sheetname):
    dfs = []
    for store in train_excels.keys():
        df = pd.read_excel(excel_paths[train_excels[store]], sheet_name=data_sheetname)
        dfs.append(df)
    df = pd.concat(dfs, axis=0)
    df = df.rename(columns={"Dealership": "store"})
    return df


def get_df_R(df):

    df_R = df[["Amount", "Category", "Type", "Source", "Department", "store"]].copy()
    df_R["Name + Description + Notes"] = pd.NA
    df_R["Notes"] = pd.NA

    df_R = df_R[
        [
            "Amount",
            "Category",
            "Type",
            "Source",
            "Department",
            "Name + Description + Notes",
            "Notes",
            "store",
        ]
    ]
    df_R.columns = [
        "Amount",
        "Category",
        "Type",
        "Source",
        "Department",
        "Name + Description + Notes",
        "Notes",
        "Dealership",
    ]
    # df_R['Source']=df['Source'].str[0:3]
    return df_R


def find_dept(acc_series, dept_series):
    try:
        acc_series = acc_series.astype(int).astype(str)
    except ValueError:
        acc_series = acc_series.astype(str)
    dept_series = acc_series.apply(lambda x: x[-1])
    replace = {
        "1": "1_NEW",
        "2": "2_USED",
        "3": "3_PARTS",
        "4": "4_SERVICE",
        "5": "4_SERVICE",
        "A": "1_NEW",
        "B": "2_USED",
        "C": "3_PARTS",
        "D": "4_SERVICE",
    }
    dept_series = dept_series.replace(replace)
    return dept_series


def rearrange_order_df_L(df_L):
    if not "Vendor Number" in df_L.columns:
        df_L["Vendor Number"] = df_L["Control#"]
    cols = [
        "Source",
        "Reference#",
        "Post Date",
        "Port",
        "Control#",
        "Debit Amount",
        "Amount",
        "Control Description",
        "Vendor Number",
        "Count",
        "Description",
        "Account",
        "Account Description",
        "Department",
    ]
    for col in cols:
        if not col in df_L.columns:
            df_L[col] = np.nan
    df_L["Control Description"] = df_L["Control Description"].fillna(
        df_L["Vendor Name"]
    )
    df_L = df_L[cols].copy()

    df_L.columns = [
        "Source",
        "Reference#",
        "Post Date",
        "Port",
        "Control#",
        "Debit Amount",
        "Credit Amount",
        "Vendor Name",
        "Vendor Number",
        "Count",
        "Description",
        "Account",
        "Account Description",
        "Department",
    ]

    return df_L


def check_duplicate(output_path, df_L, data_sheetname):

    df_out = pd.read_excel(output_path, sheet_name=data_sheetname)

    df_out["Post Date"] = pd.to_datetime(df_out["Post Date"])
    data_month = df_L["Post Date"].dt.month.unique()
    data_year = df_L["Post Date"].dt.year.unique()

    for y in data_year:
        for m in data_month:
            mask = (df_out["Post Date"].dt.month == m) & (
                df_out["Post Date"].dt.year == y
            )
            duplicate = df_out.loc[mask]
            if len(duplicate) > 0:
                duplicate = True
                break
            else:
                duplicate = False
    return duplicate


def new_vendor_cate(new_vendor, df_L, df_R):
    vendor_type = {}
    vendor_cate = {}
    for vendor in new_vendor.keys():
        # mask=clean_vendor(df_L)==vendor
        mask = df_L["Vendor Name"] == vendor
        vendor_cate[vendor] = df_R.loc[mask, "Category"].unique()
        vendor_type[vendor] = df_R.loc[mask, "Type"].unique()
    return vendor_cate, vendor_type


def lemmatize_text(text):
    if not os.path.exists(os.path.join(nltk.data.find("corpora"), "wordnet.zip")):
        nltk.download("wordnet")
        nltk.download("punkt")

    lemmatizer = WordNetLemmatizer()
    tokens = word_tokenize(text)
    lemmatized_tokens = [lemmatizer.lemmatize(token) for token in tokens]
    return " ".join(lemmatized_tokens)


def clean_text_col(df_series):
    return_series = df_series.fillna("")
    return_series = return_series.astype(str)
    return_series = return_series.str.upper()

    ## remove months
    months = list(calendar.month_name)[1:] + list(calendar.month_abbr)[1:]
    months = [m.upper() for m in months]
    for month in months:
        return_series = return_series.str.replace(month, "")
    words = ["MONTH", "MONTHLY", "MNT", "ACCRUAL"]
    for word in words:
        return_series = return_series.str.replace(word, "")

    ## remove anything other than letters
    def clean_text(text):
        _text = re.sub(r"[^a-zA-Z]", " ", text)
        text = re.sub(r"\s+", " ", _text).strip()
        return text

    return_series = return_series.apply(clean_text)
    return return_series


def get_textfeature_tfidf(text_col):
    text_features = text_col.astype(str).apply(lambda x: x.strip()).fillna("")

    _vocab = np.array([])
    for v in text_features.unique():
        v_list = v.split()
        v_lower = []
        for v in v_list:
            if type(v) == str:
                v_lower.append(v.lower())
            else:
                v_lower.append(v)
        v = np.array(v_lower)
        _vocab = np.append(_vocab, v)
    vocabs = np.unique(_vocab)

    text_vectorizer = TfidfVectorizer(
        vocabulary=vocabs,
        ngram_range=(1, 3),
        analyzer="word",
        encoding="utf-8",
        decode_error="replace",
    )
    text_vectorizer = text_vectorizer.fit(vocabs)
    text_features_vectorized = text_vectorizer.transform(text_features)

    df_textfeature = pd.DataFrame(text_features_vectorized.toarray())

    return df_textfeature, text_vectorizer


def get_labels(df, label_cols: list):
    encoded_labels = {}
    encoder = {}
    # unique_values={}

    for col in label_cols:
        unique_values = np.append(df[col].unique(), "MISCELLANEOUS")
        label_encdoer = LabelEncoder().fit(unique_values)
        # label_encdoer=OrdinalEncoder(handle_unknown='use_encoded_value', unknown_value=-1)
        encoded_label = label_encdoer.transform((df[col].values))
        # encoded_label = label_encdoer.transform((df[col].values).reshape(-1,1))
        # output
        # label_dict=dict(zip([*range(0,len(encoded_label))],encoded_label))

        # encoded_labels[col]=label_dict
        encoded_labels[col] = encoded_label
        encoder[col] = label_encdoer
    # train_encoded_labels=encoded_labels.copy()
    df_encoded_labels = pd.DataFrame(encoded_labels)
    return df_encoded_labels, encoder


def get_labels_v2(df):
    label_features = df[
        ["Vendor Name", "Source #", "Account", "store", "Reference#", "Control#"]
    ]
    label_cols = label_features.columns
    encoded_labels = {}
    encoder = {}
    # unique_values={}

    for col in label_cols:
        unique_values = np.append(label_features[col].unique(), "MISCELLANEOUS")
        label_encdoer = LabelEncoder().fit(unique_values)
        # label_encdoer=OrdinalEncoder(handle_unknown='use_encoded_value', unknown_value=-1)
        encoded_label = label_encdoer.transform((label_features[col].values))
        # encoded_label = label_encdoer.transform((df[col].values).reshape(-1,1))
        # output
        # label_dict=dict(zip([*range(0,len(encoded_label))],encoded_label))

        # encoded_labels[col]=label_dict
        encoded_labels[col] = encoded_label
        encoder[col] = label_encdoer
    # train_encoded_labels=encoded_labels.copy()
    df_encoded_labels = pd.DataFrame(encoded_labels)
    return df_encoded_labels.values


def filter_minority(df, minority=4):
    if minority == 0:
        return df

    _df = df.copy()

    _df["cnt"] = _df.Category
    replace = _df["Category"].value_counts().to_dict()
    _df.cnt = _df.cnt.replace(replace)

    mask = _df.cnt > minority
    _df = _df.loc[mask]
    _df.pop("cnt")
    _df = _df.reset_index(drop=True)
    return _df


def Type_col_mapping(cate_series, cate_type_map, default_Type_value):
    type_series = cate_series.map(cate_type_map)
    type_series = type_series.fillna(default_Type_value)
    return type_series


"""
ARCHIVE
def get_cate_type_dict(train_df,cate_type_dict={},change={}):
    types=train_df['Type'].unique()
    types_option={}
    for v in types:
        if not pd.isna(v):
            types_option[v[0].lower()]=v
    for cate in train_df.Category.unique():
        mask=train_df.Category==cate
        type=train_df.loc[mask,'Type'].unique()
        if len(type)>1:
            # type=input('Please Enter the correct Type of [{}],\nPossible Types:{}  :'.format(cate,type))
            text='Please Enter the correct Type of [{}],\nPossible Types:{}  :'.format(cate,type)
            task=types_option
            type=inputv2(text,task)
            change[cate]=type
        cate_type_dict[cate]=type[0]
    return cate_type_dict,change

def correct_type(excel_paths,train_excels,cate_type_change,data_sheetname,confirm_to_run=False):
    if (confirm_to_run) & (len(cate_type_change)>0):
        for key in train_excels.keys():
            xl=excel_paths[train_excels[key]]

            _df=pd.read_excel(xl,sheet_name=data_sheetname)
            for cate in cate_type_change.keys():
                mask=_df.Category==cate
                _df.loc[mask,'Type']=cate_type_change[cate]

            paste_df_to_xl(xl,data_sheetname,'A2',_df)
            # wb=xw.Book(xl)
            # api=xw.apps.active
            # sht=wb.sheets[data_sheetname]

            # sht['A2'].value=_df.values
            # # sht['A2'].options(pd.DataFrame,index=False,header=False,expand='table').value=_df
            # wb.save()
            # api.quit()
"""


def get_unique_yearmonths(df, date_col_name, date_format="%Y-%m-%d"):
    df[date_col_name] = pd.to_datetime(df[date_col_name], format=date_format)

    yearmonths = []
    year = df[date_col_name].dt.year.unique()
    for y in year:
        year_mask = df[date_col_name].dt.year == y
        months = df.loc[year_mask, date_col_name].dt.month.unique()
        for m in months:
            yearmonths.append(str(y) + str(m).zfill(2))
    return yearmonths


def remove_by_yearmonth(output_path, data_sheetname, date_col_name, yearmonth_list):
    """
    same func as in corr_xl_func.py
    yearmonth_list : [yyyymm]
    """
    df = pd.read_excel(output_path, sheet_name=data_sheetname)
    df[date_col_name] = pd.to_datetime(df[date_col_name])

    for yearmonth in yearmonth_list:
        year = int(yearmonth[0:4])
        month = int(yearmonth[4:])
        mask = (df[date_col_name].dt.year == year) & (
            df[date_col_name].dt.month == month
        )
        df = df.loc[~mask]

    wb = load_workbook(output_path)
    ws = wb[data_sheetname]

    ws.delete_rows(2, ws.max_row)
    wb.save(output_path)
    wb.close()
    paste_df_to_xl(output_path, data_sheetname, "A2", df)


def append_data_openpyxl(store, df_L, df_R, output_paths, yearmonth, data_sheetname):
    path = output_paths[store]
    # Read Excel file to determine the max row
    # df_out = pd.read_excel(path, sheet_name=data_sheetname)
    # maxrow = len(df_out) + 2

    # Load workbook and select sheet
    # wb = load_workbook(path)
    # sht = wb[data_sheetname]

    _df = pd.concat([df_L, df_R], axis=1)
    paste_df_to_xl(path, data_sheetname, "max", _df)

    # Output information
    output_name = path.split("/")[-1]
    print(
        f"[{len(df_L)}] rows of data on [{yearmonth}] are appended to [{output_name}]"
    )

    return path


def replace_data(store, df_L, df_R, output_paths, yearmonth, data_sheetname):
    xl_path = output_paths[store]
    date_col_name = "Post Date"
    ### get assigned yearmonth
    yearmonth_list = get_unique_yearmonths(df_L, date_col_name, date_format="%Y-%m-%d")

    ### read df
    df = pd.read_excel(xl_path, sheet_name=data_sheetname)
    df[date_col_name] = pd.to_datetime(df[date_col_name])
    ### remove assigned yearmonth
    for yearmonth in yearmonth_list:
        year = int(yearmonth[0:4])
        month = int(yearmonth[4:])
        mask = (df[date_col_name].dt.year == year) & (
            df[date_col_name].dt.month == month
        )
        df = df.loc[~mask]
        df = df.where(pd.notnull(df), None)
    ### load wb
    wb = load_workbook(xl_path)
    ws = wb[data_sheetname]

    # Clear all filters
    if is_any_data_filtered(ws):
        ws.auto_filter.filterColumn.clear()
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].hidden = False
    ### remove all data
    ws.delete_rows(2, ws.max_row)

    cell_name = "A2"

    cell_letter_idx = int(
        column_index_from_string(re.search(r"\D+", cell_name).group())
    )
    cell_number = re.search(r"\d+", cell_name).group()

    ### write ORG data , after dropping assigned yearmonth
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), int(cell_number)
    ):
        for c_idx, value in enumerate(row, cell_letter_idx):  # Columns start at 1
            ws.cell(row=r_idx, column=c_idx, value=value)

    ### get new data for assigned yearmonth
    _df = pd.concat([df_L, df_R], axis=1)
    _df = _df.where(pd.notnull(_df), None)

    cell_name = "MAX"
    cell_letter_idx = column_index_from_string("A")
    cell_number = ws.max_row + 1

    # Write df_L df_R on max cells
    for r_idx, row in enumerate(
        dataframe_to_rows(_df, index=False, header=False), int(cell_number)
    ):
        for c_idx, value in enumerate(row, cell_letter_idx):  # Columns start at 1
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(xl_path)
    wb.close()

    return xl_path


def get_index(output_path, data_sheetname):
    df = pd.read_excel(output_path, data_sheetname)
    position = len(df.columns) - 1
    df = df.drop(columns=["Index"])
    df = df.reset_index()
    index_col = df.pop("index")
    df.insert(position, "Index", index_col)
    return df


def get_cate_map(df, clf_option, enable_col_mapping):
    """
    Generate a category mapping from a DataFrame based on a specified column.

    Parameters:
    - df: DataFrame containing the data.
    - clf_option: Classifier option for handling duplicates ('simple_map' or other).
    - enable_col_mapping: Column to use for mapping categories.

    Returns:
    - cate_map: A dictionary mapping unique values in the specified column to categories.
    - info: Information message about any duplicate index values encountered.
    """

    col = enable_col_mapping
    df[col] = df[col].astype(str)
    _df = df[[col, "Category"]].copy()
    _df = _df.drop_duplicates()
    map_series = _df.set_index([col])["Category"]
    col_duplicated_bools = map_series.index.duplicated()

    if not col_duplicated_bools.any():
        # * No duplicates found, proceed
        info = None
        cate_map = map_series.to_dict()
        return cate_map, info
    else:
        duplicated_col_values = map_series.loc[col_duplicated_bools].index.tolist()
        duplicated_df = _df.loc[_df[col].isin(duplicated_col_values)]
        non_duplicated_df = _df.loc[~_df[col].isin(duplicated_col_values)]

        if clf_option == "simple_map":
            info = f"ERROR: __acc_replace__ - Duplicated Index Found in Train_df for Simple Mapping\n {duplicated_df}"
            cate_map = map_series.to_dict()
            return cate_map, info
        else:
            # * For other classifiers, drop duplicate index
            info = None
            map_series = non_duplicated_df.set_index([col])["Category"]
            cate_map = map_series.to_dict()
            return cate_map, info


def split_by_col_cate_dict(df, col, col_cate_dict):
    df[col] = df[col].astype(str)
    mask = df[col].isin(col_cate_dict.keys())
    df_match = df.loc[mask].copy()
    df_clf = df.loc[~mask].copy()
    ### df_match Cate result from mapping dict
    ### df_clf : assign back to df, going to go through normal CLF process
    return df_clf, df_match


def join_text_columns(df: pd.DataFrame, base_col: str, extra_cols: list):
    text_col = df.copy()[base_col]
    if len(extra_cols) > 0:
        for col in extra_cols:
            text_col = text_col + " " + df[col].astype(str).fillna("")
    text_col = text_col.str.strip()
    text_col = text_col.fillna("")
    return text_col


def cleanse_col(df):
    _df=df.copy()
    if "Category" in _df.columns:
        _df["Category"] = _df["Category"].str.strip()

    _df["Source Name"] = _df["Source Name"].apply(int_to_str)
    _df["Source #"] = _df["Source #"].apply(int_to_str)
    _df["Account"] = _df["Account"].apply(int_to_str)
    _df["Reference#"] = clean_text_col(_df["Reference#"]).str.upper()
    _df["Control#"] = clean_text_col(_df["Control#"]).str.upper()
    _df["Description"] = _df["Description"].fillna(_df["Control#"])
    _df["Description"] = clean_text_col(_df["Description"])
    _df["Description"] = _df["Description"].apply(lambda x: lemmatize_text(x))
    _df["Vendor Name"] = clean_text_col(_df["Vendor Name"])
    _df["Vendor Name"] = _df["Vendor Name"].fillna("").astype(str).str.upper()
    return _df


def get_feature(df, x_encoder: dict, label_cols: list, text_vectorizer=None):
    if len(x_encoder) == 0:
        float_features = df["Amount"].fillna(0).reset_index(drop=True)
        date_features = pd.to_datetime(df["Post Date"])

        text_col = join_text_columns(df, "Description", [])
        df_textfeature, text_vectorizer = get_textfeature_tfidf(text_col)
        df_encoded_labels, x_encoder = get_labels(df, label_cols)
        feature = df_textfeature.join(df_encoded_labels)

        return feature, x_encoder, text_vectorizer
    else:
        float_features = df["Amount"].fillna(0)
        date_features = pd.to_datetime(df["Post Date"])

        text_col = join_text_columns(df, "Description", [])
        text_feature = text_vectorizer.transform(text_col).toarray()
        df_textfeature = pd.DataFrame(text_feature)

        encoded_labels = {}
        for col in label_cols:

            mask = df[col].isin(x_encoder[col].classes_)
            # handle unseen values
            df.loc[~mask, col] = "MISCELLANEOUS"
            encoded_labels[col] = x_encoder[col].transform(df[col])

        df_encoded_labels = pd.DataFrame(encoded_labels)

        feature = df_textfeature.join(df_encoded_labels)
        feature.columns = feature.columns.astype(str)

        return feature


def get_y(df):
    y_encoder = LabelEncoder().fit(df["Category"])
    y = y_encoder.transform(df["Category"])
    y = pd.DataFrame(y, columns=["y_label"]).y_label

    return y, y_encoder


class txtColumnSelector(BaseEstimator, TransformerMixin):
    def __init__(self, base_col=None, extra_cols=None):
        self.base_col = base_col
        self.extra_cols = extra_cols

    def fit(self, X, y=None):
        return self

    def transform(self, X):
        if self.base_col is None or self.extra_cols is None:
            return X
        return join_text_columns(X, self.base_col, self.extra_cols)


class xbg_FeatureSelector(BaseEstimator, TransformerMixin):
    def __init__(self, params=None):
        self.params = params
        # self.num_class=num_class
        # self.params = params
        # self.threshold = threshold
        self.model = None

    def fit(self, X, y):
        # Fit the XGBoost model
        self.model = xgb.XGBClassifier(**self.params)
        self.model.fit(X, y)
        return self

    def transform(self, X):
        # Calculate feature importances
        feature_importances = self.model.feature_importances_
        self.threshold = np.sort(np.unique(feature_importances))[1]

        # Select features based on the threshold
        selected_features = X[:, feature_importances >= self.threshold]
        return selected_features


class ZeroArrayTransformer(BaseEstimator, TransformerMixin):
    def __init__(self):
        pass

    def fit(self, X, y=None):
        # This method is not needed for this transformer, so it's just a placeholder
        return self

    def transform(self, X):
        # Assuming X is a DataFrame, convert it to a NumPy array if necessary
        if isinstance(X, pd.DataFrame):
            X = X.to_numpy()

        # Create an array of zeros with the same number of rows as X
        zero_array = np.zeros((X.shape[0], 1))
        return zero_array


def save_cate_type_map(train_df):
    path = "function/Cate_Type_map.csv"
    cate_type_df = train_df.set_index(["Category"])["Type"]
    cate_type_df = cate_type_df.reset_index()
    cate_type_df = cate_type_df.drop_duplicates()
    if not os.path.exists(path):
        cate_type_df.to_csv(path, index=False)
    else:
        ### load map
        old_map = pd.read_csv(path)
        new_map = cate_type_df.set_index(["Category"])["Type"].to_dict()
        type = old_map["Category"].map(new_map)
        old_map["Type"] = type.fillna(old_map.Type)
        result = pd.concat([old_map, cate_type_df])
        result = result.drop_duplicates()
        result.to_csv(path, index=False)
