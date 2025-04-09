import os

# import shutil
import re
import pandas as pd
import numpy as np
from datetime import datetime

# import sys
# import threading
# # import calendar
import yaml

# # import pythoncom
# import win32com.client
import json
import zipfile
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
import warnings

warnings.filterwarnings(action="ignore", category=FutureWarning)


def add_mod_time_suffix(filepath, modified_time=None):
    if not modified_time:
        modified_time = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime(
            "%Y%m%d_%H%M%S"
        )

    dir = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    ext = filename.split(".")[-1]
    name = filename.split(".")[0]
    newname = name + "-" + modified_time + "." + ext
    return os.path.join(dir, newname)


def locate_excel(path, excel_path):
    excels = [
        file
        for file in os.listdir(path)
        if (not file.startswith("~$")) and (file.endswith(".xlsx"))
    ]
    new_excel_path = {
        file: os.path.join(path, file)
        for file in os.listdir(path)
        if (not file.startswith("~$")) and (file.endswith(".xlsx"))
    }
    excel_path.update(new_excel_path)
    return excels, excel_path


def locate_folders(folder_prefix, rawdata_path, keyword=""):
    # v2.0 20230622 locate in upper layer
    folders = []
    PATHs = []
    # path=os.getcwd()
    # rawdata_path='/raw_data/'
    # if upper_layer:
    # rawdata_path='../raw_data/'
    _path = os.path.join(rawdata_path, folder_prefix)
    for folder in os.listdir(_path):
        if folder.startswith(keyword):
            folders.append(folder)
            PATH = os.path.join(_path, folder)
            PATHs.append(PATH)

    return folders, PATHs


def latest_month(folders, paths):
    # v1.0
    yearmonth = []
    target = {}
    for folder, path in zip(folders, paths):
        # v=int(re.findall(r'/d+',folder))
        v = int(re.findall(r"\d+", folder)[0])
        yearmonth.append(v)
        target[v] = path
    v = max(yearmonth)
    return target[v], v


def select_month(folders, paths, yearmonth):
    # v1.0
    # yearmonth=[]
    target = {}
    for folder, path in zip(folders, paths):
        if re.search(yearmonth, folder):
            v = folder
            target[v] = path
        # v=int(re.findall(r'/d+',folder))
        # v=int(re.findall(r'\d+',folder)[0])
        # yearmonth.append(v)
        # target[v]=path
    # v=max(yearmonth)
    folder_name = v
    return target[v], folder_name


def keyin_argv_store(argv, client_info, location_code):
    try:
        argv = eval(argv)
        if type(argv) == list:
            return argv
        else:
            load_from_client_info = True
    except NameError:
        load_from_client_info = True
    # if load_from_client_info:
    if location_code == argv:
        assigned_store = client_info["location"][location_code]
    elif argv in client_info["location"][location_code].values():
        assigned_store = [argv]
    else:
        assigned_store = client_info["location"][location_code]

    if not client_info["sub_stores"]:
        return assigned_store
    else:
        for store in assigned_store:
            if store in client_info["sub_stores"].keys():
                sub_stores = client_info["sub_stores"][store]
                assigned_store = assigned_store + sub_stores
                assigned_store = list(set(assigned_store))
        return assigned_store


class StoreAssigner:
    def __init__(self, client_info, location_code):
        self.client_info = client_info
        self.location_code = location_code

    def assign_store(self, argv):
        try:
            argv = eval(argv)
            if type(argv) == list:

                return [v.upper() for v in argv]
            else:
                argv = argv.upper()
                load_from_client_info = True
        except NameError:
            argv = argv.upper()
            load_from_client_info = True
        if load_from_client_info:

            if self.location_code == argv:
                assigned_store = self.client_info["location"][self.location_code]
            elif argv in self.client_info["location"][self.location_code]:
                assigned_store = [argv]
            else:
                assigned_store = self.client_info["location"][self.location_code]

            if not self.client_info["sub_stores"]:
                return assigned_store
            else:
                for store in assigned_store:
                    if store in self.client_info["sub_stores"].keys():
                        sub_stores = self.client_info["sub_stores"][store]
                        assigned_store = assigned_store + sub_stores
                        assigned_store = list(set(assigned_store))

                return assigned_store


def keyin_argv_yearmonth(argv):
    if re.search(r"\d{6}", argv):
        assigned_yearmonth = argv
    else:
        while True:
            option = input(
                "wrong arguement. Please enter YearMonth ie: 202301\n"
            ).upper()
            if option in ["Q", "QUIT"]:
                quit()
            elif re.search(r"\d{6}", option):
                assigned_yearmonth = option
                break
    return assigned_yearmonth


def inputv2(text="", task={"y": True, "n": False}):
    while True:
        option = input(text).lower()
        if option in task.keys():
            break
        else:
            print("Invalid option. Please Re-Enter.")
    return task[option]


def concat_df_indict(_dict, n=0):
    # v1.0
    cnt = 0
    for key in _dict.keys():
        if cnt == 0:
            df = _dict[key]
            cnt = cnt + 1
        else:
            df2 = _dict[key]
            df = pd.concat([df, df2], axis=n)
    return df


def castdate(df, col, just_date=True):
    # Convert column to int64 to handle timestamps correctly
    series = df[col].copy()
    if series.isna().sum() > 0:
        series = series.dropna()

    series = pd.to_numeric(series, errors="coerce").astype("Int64")
    series = series.replace(0, np.nan)

    # Use vectorized operation to convert timestamps to dates
    series = series.map(
        lambda x: (
            datetime.fromtimestamp(x / 1000) if isinstance(x, int) else "str error"
        ),
        na_action="ignore",
    )

    if just_date:
        series = series.apply(lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else "")
        return series
    else:
        return series


# def input_with_timeout(prompt, default_value, timeout=5):
#     print(prompt, end='')
#     sys.stdout.flush()

#     # Create an event to signal input received
#     input_event = threading.Event()

#     # Define a function to be called on timeout
#     def timeout_handler():
#         input_event.set()

#     # Start the timer thread
#     timer_thread = threading.Timer(timeout, timeout_handler)

#     timer_thread.start()
#     # Read input in a separate thread
#     input_text = None

#     def input_thread_func():
#         nonlocal input_text
#         input_text = sys.stdin.readline().strip()
#         input_event.set()

#     input_thread = threading.Thread(target=input_thread_func)
#     input_thread.start()

#     # Wait for input or timeout
#     input_event.wait()

#     # Cancel the timer thread
#     timer_thread.cancel()

#     if input_text:
#         return input_text
#     else:
#         # If no input is received, return the default value
#         print(f"\nNo input received. Using default value: {default_value}")
#         return default_value

# def select_yearmonth_folder_timeout(PATH):
#     folders=os.listdir(PATH)
#     folders.sort(reverse=True)

#     # Usage example
#     no = input_with_timeout("Select folder (in 5 sec):\n"+','.join(folders)+'\n', "0")
#     return folders[int(no)]


def load_config(file_path="config.yml"):
    if not os.path.exists(file_path):
        folder = os.path.dirname(file_path)
        os.makedirs(folder, exist_ok=True)
        with open(file_path, "w") as file:
            file.write("")

    with open(file_path, "r") as file:
        config = yaml.load(file, Loader=yaml.SafeLoader)
    return config


def int_to_str(x):

    if pd.isnull(x):  # Check for NaN values
        return x
    elif x is None:
        return x
    try:
        return str(int(float(x)))
    except (ValueError, TypeError):
        return x
    except OverflowError:
        # * float is too large to convert
        return x


# def create_shortcut(excel_file_path, shortcut_dir):
#     shortcut_dir=shortcut_dir.replace('/','\\')
#     excel_file_path=os.path.abspath(excel_file_path)

#     shell = win32com.client.Dispatch("WScript.Shell")
#     shortcut = shell.CreateShortCut(os.path.join(shortcut_dir, os.path.basename(excel_file_path) + '.lnk'))
#     shortcut.Targetpath = excel_file_path
#     shortcut.save()


def setup_col_table(df, shtname):
    cols = df.columns
    new_cols = []
    for col in cols:
        col = col.replace(" ", "_").lower()
        new_cols.append(col)
    col_df = pd.DataFrame([new_cols], columns=list(cols))
    col_df.to_excel("Col_table.xlsx", index=False, sheet_name=shtname)


def get_df_from_coltbs(coltbl_path, file_path):
    filename = os.path.basename(file_path)
    if file_path.endswith(".xlsx"):
        df = pd.read_excel(filename)
    else:
        df = pd.read_csv(file_path)
    col_df = pd.read_excel(coltbl_path, sheet_name=filename)
    col_df = col_df.dropna(axis=1)
    for col in df.columns:
        if col not in col_df.columns:
            df.pop(col)
    df.columns = col_df.loc[0]
    return df


def load_col_dfs(coltbl_path):
    col_dfs = {
        sht: pd.read_excel(coltbl_path, sheet_name=sht)
        for sht in get_excel_sheet_names(coltbl_path)
    }
    return col_dfs


def fix_col_by_coltbl(df, coltbl: str, sht_name: str):
    """
    coltbl :str -> coltbl_path
    coltbl :dict -> dict contain all df_col in coltbl
    """
    df = df.copy()

    if type(coltbl) == str:
        col_df = pd.read_excel(coltbl, sheet_name=sht_name)
    elif type(coltbl) == dict:
        col_df = coltbl[sht_name]

    col_df = col_df.dropna(axis=1, how="all")

    for col in df.columns:
        if not col in col_df.columns:
            df.pop(col)
    for n in range(1, len(col_df)):
        dup_col_df = pd.DataFrame([col_df.loc[n]])
        dup_col_df = dup_col_df.dropna(axis=1)
        for old_col, add_col in zip(dup_col_df.columns, dup_col_df.values[0]):
            df[add_col] = df[old_col]
        col_df = col_df.drop(index=n)
    rename_map = dict(zip(col_df.columns.to_list(), col_df.values[0]))
    df = df.rename(columns=rename_map)
    for col in rename_map.values():
        if not col in df.columns:
            df[col] = pd.NA
    return df


def get_colseq(df, coltbl: str, sht_name: str):
    if type(coltbl) == str:
        colseq = pd.read_excel(coltbl, sheet_name=sht_name)
    elif type(coltbl) == dict:
        colseq = coltbl[sht_name]
    cols = colseq.columns.to_list()
    df = df[cols].copy()
    return df


def fix_dtype_by_tbl(df, dtypetbl_path, sht_name):
    df = df.copy()
    df_dtype = pd.read_excel(dtypetbl_path, sheet_name=sht_name)
    df_dtype = df_dtype.dropna(axis=1)
    data = {"cols": df_dtype.columns.to_list(), "dtypes": df_dtype.loc[0].values}
    df_dtype = pd.DataFrame(data)

    mask = df_dtype["dtypes"].isin(["str", "int", "float"])
    for col, d in zip(df_dtype.loc[mask, "cols"], df_dtype.loc[mask, "dtypes"]):
        mask = df[col].isna()
        df.loc[~mask, col] = df.loc[~mask, col].astype(d)

    mask = df_dtype["dtypes"] == "int_to_str"
    for col in df_dtype.loc[mask, "cols"]:
        df[col] = df[col].apply(int_to_str)

    mask = df_dtype["dtypes"] == "date"
    for col in df_dtype.loc[mask, "cols"]:
        df[col] = pd.to_datetime(df[col]).dt.strftime("%Y-%m-%d")

    mask = df_dtype["dtypes"] == "time"
    for col in df_dtype.loc[mask, "cols"]:
        df[col] = pd.to_datetime(df[col]).dt.strftime("%H:%M:%S")

    mask = df_dtype["dtypes"] == "datetime"
    for col in df_dtype.loc[mask, "cols"]:
        df[col] = pd.to_datetime(df[col])

    return df


def load_json_data(output):
    # Split the output into lines
    lines = output.split("\n")

    # Look for the line containing the JSON data and parse it
    json_data = None
    for line in lines:
        if line.strip():
            try:
                json_data = json.loads(line)
                break  # Stop searching when the JSON is found
            except json.JSONDecodeError:
                pass  # Ignore non-JSON lines
    return json_data


def get_excel_sheet_names(file_path):
    sheets = []
    with zipfile.ZipFile(file_path, "r") as zip_ref:
        xml = zip_ref.read("xl/workbook.xml").decode("utf-8")
    for s_tag in re.findall("<sheet [^>]*", xml):
        sheets.append(re.search('name="[^"]*', s_tag).group(0)[6:])
    return sheets


def parse_argv(argvs, func_list_when_True, otherwise_if_fasle):

    while len(argvs) <= (len(func_list_when_True)):
        argvs.append(False)
    if len(argvs) > len(func_list_when_True):
        """If provided argvs MORE than function_list, Return the apply Function on the Pasred_argvs + extra argv(Without changing anything)"""
        parsing_argv = argvs[1 : len(func_list_when_True) + 1]
        extra_argvs = argvs[len(func_list_when_True) + 1 :]
    else:
        parsing_argv = argvs[1:]
        extra_argvs = []
    parsed_argv = [argvs[0]]
    for n, argv in enumerate(parsing_argv):
        if argv:
            if func_list_when_True[n]:
                parsed_argv.append(func_list_when_True[n](argv))
            else:
                parsed_argv.append(argv)
        else:
            parsed_argv.append(otherwise_if_fasle[n])
    return parsed_argv + extra_argvs


def is_any_data_filtered(sht):
    """
    Check if there is any active filter in the given worksht.
    Returns True if any filter is active, otherwise False.
    """
    if sht.auto_filter.ref is None:
        return False

    else:
        if sht.auto_filter.filterColumn:
            return True
        else:
            return False


def clear_xl_filter(xl_path, sht_name):
    try:
        wb = load_workbook(xl_path)
    except KeyError:
        wb = load_workbook(filename=xl_path, data_only=True, keep_links=False)
    sht = wb[sht_name]

    # Clear all filters
    if is_any_data_filtered(sht):
        sht.auto_filter.filterColumn.clear()
        for i in range(2, sht.max_row + 1):
            sht.row_dimensions[i].hidden = False
    wb.save(xl_path)
    wb.close()


def paste_df_to_xl(xl_path, sht_name, cell_name, df):
    try:
        wb = load_workbook(xl_path)
    except KeyError:
        wb = load_workbook(filename=xl_path, data_only=True, keep_links=False)

    sht = wb[sht_name]

    df = df.where(pd.notnull(df), None)

    # Clear all filters
    if is_any_data_filtered(sht):
        sht.auto_filter.filterColumn.clear()
        for i in range(2, sht.max_row + 1):
            sht.row_dimensions[i].hidden = False

    if cell_name.upper() == "MAX":
        cell_letter_idx = column_index_from_string("A")
        cell_number = sht.max_row + 1
    else:
        cell_letter_idx = int(
            column_index_from_string(re.search(r"\D+", cell_name).group())
        )
        cell_number = re.search(r"\d+", cell_name).group()

    # Write df_L values to the appropriate cells
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), int(cell_number)
    ):
        for c_idx, value in enumerate(row, cell_letter_idx):  # Columns start at 1
            sht.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    wb.save(xl_path)
    wb.close()


def keyin_truefalse_argv(argv):
    task = {
        "y": True,
        "n": False,
        "True": True,
        "False": False,
        True: True,
        False: False,
    }
    return task[argv]


def makedir_for_file(file_path):
    if re.search(r".*\.[\w]+$", os.path.basename(file_path)):
        isfolder = False
    else:
        isfolder = True
    if isfolder:
        dir_path = file_path
    else:
        dir_path = os.path.dirname(file_path)
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)


def select_file(extension=False, search_by=False, dir=os.getcwd(), auto_select=True):
    if extension:
        files = [file for file in os.listdir(dir) if file.endswith(extension)]
    else:
        files = os.listdir(dir)
    if search_by:
        files = [file for file in files if re.search(search_by, file)]
    sorted_files = sorted(files, key=lambda x: os.path.getmtime(x), reverse=True)
    if auto_select:
        excel = sorted_files[0]
    else:
        selected_file = input("Select :\n{}".format(" , ".join(sorted_files)) + "\n")
        excel = sorted_files[int(selected_file)]
    return excel


def combine_col(df: pd.DataFrame, base_col: str, extra_col: str):
    mask = pd.isna(df[extra_col])
    _df = df.loc[~mask].copy()
    if len(_df) == 0:
        # If all values in extra_col are NaN, return NaN series
        return pd.Series([np.nan] * len(df))

    _df[extra_col] = _df[extra_col].apply(int_to_str)
    try:
        _df[extra_col] = _df[extra_col].str.strip()
    except AttributeError:
        _df[extra_col] = _df[extra_col]
    text_col = _df[base_col] + "|" + _df[extra_col].str.strip()
    text_col = text_col.str.strip()

    # Create a series with NaN for rows where extra_col is NaN
    result = pd.Series(index=df.index)
    result.loc[~mask] = text_col
    result.loc[mask] = df.loc[mask, base_col]

    # Remove trailing separators and check if result equals base_col
    result = result.str.rstrip("|")
    base_mask = result == df[base_col]
    result[base_mask] = np.nan

    return result
