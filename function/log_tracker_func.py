import os
import pandas as pd
from datetime import datetime
import re
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import coordinate_to_tuple
from .python_onedrive_func import *
import warnings

warnings.filterwarnings(action="ignore", category=pd.errors.DtypeWarning)
warnings.filterwarnings(action="ignore", category=UserWarning)


# from utilities import select_file
def select_file(extention=False, search_by=False, auto_select=False, dir=None):
    if dir is None:
        files = os.listdir()
    else:
        files = [os.path.join(dir, file) for file in os.listdir(dir)]
    if extention:
        _files = []
        for file in files:
            if file.endswith(extention):
                if search_by:
                    if re.search(search_by, file):
                        _files.append(file)
                else:
                    _files.append(file)
        files = _files
    sorted_files = sorted(files, key=lambda x: os.path.getmtime(x), reverse=True)
    if auto_select:
        excel = sorted_files[0]
    else:
        selected_file = input(
            "Select Analysis to be corrected:\n{}".format(" , ".join(sorted_files))
            + "\n"
        )
        excel = sorted_files[int(selected_file)]
    return excel


def convert_str(var, global_):
    for name in global_.keys():
        if name is var:
            return name


# def get_value_in_global(var, global_):
#     """get key if values match the var"""
#     try:
#         return [name for name, value in global_.items() if value is var][0]
#     except IndexError:
#         return None


def create_substituted_dict(subs_dict, name_path_dict, search_by=None):
    result_dict = {}

    # Get all possible combinations
    for key, pattern in name_path_dict.items():
        # Initialize with first key-value pair
        result_dict[key] = pattern

        # Find all substitution patterns in the key and value
        for sub_key, sub_values in subs_dict.items():
            sub_pattern = "{" + sub_key + "}"
            temp_dict = {}

            # Only process if the pattern exists in key or value
            if sub_pattern in key or sub_pattern in pattern:
                # For each existing entry, create new entries with substitutions
                for existing_key, existing_value in result_dict.items():
                    # If search_by is provided
                    if search_by:
                        if search_by in existing_key:
                            for sub_val in sub_values:
                                new_key = existing_key.replace(sub_pattern, sub_val)
                                new_value = existing_value.replace(sub_pattern, sub_val)
                                temp_dict[new_key] = new_value
                        else:
                            temp_dict[existing_key] = existing_value
                    # For all other cases
                    else:
                        # If the current sub_values has multiple elements
                        if len(sub_values) > 1:
                            for sub_val in sub_values:
                                new_key = existing_key.replace(sub_pattern, sub_val)
                                new_value = existing_value.replace(sub_pattern, sub_val)
                                temp_dict[new_key] = new_value
                        else:
                            # Single value, just do one replacement
                            new_key = existing_key.replace(sub_pattern, sub_values[0])
                            new_value = existing_value.replace(
                                sub_pattern, sub_values[0]
                            )
                            temp_dict[new_key] = new_value

                result_dict = temp_dict

    return result_dict


def remove_temp_dict(_dict):
    _new_dict = {k: v for k, v in _dict.items() if not re.search(r"{\w+_sub}", v)}
    new_dict = {k: v for k, v in _new_dict.items() if not re.search(r"{\w+_sub}", k)}

    return new_dict


def load_file_if_dir(name_path_dict):
    for name, path in name_path_dict.items():
        if os.path.isdir(path):
            try:
                name_path_dict[name] = select_file(
                    extention=".csv",
                    search_by=datetime.now().strftime("%y-%m-01"),
                    auto_select=True,
                    dir=path,
                )
            except IndexError:
                name_path_dict[name] = "File Not Found"

    return name_path_dict


def get_updatetime(name_path_dict):
    update_time = {}
    for name, path in name_path_dict.items():
        if os.path.exists(path):
            update_time[name] = (
                f"{datetime.fromtimestamp(os.path.getmtime(path)):%Y-%m-%d %H:%M:%S}"
            )
        else:
            update_time[name] = "File Not Found"
    return update_time


def get_mod_date_from_df(date_col_dict, name_path_dict):
    max_mod_date = {}
    for name, path in name_path_dict.items():
        if not os.path.exists(path):
            max_mod_date[name] = "File Not Found"
        else:
            for suffix in date_col_dict:
                if name.startswith(suffix):
                    col = date_col_dict[suffix]
                    df = pd.read_csv(path, parse_dates=[col])
                    max_mod_date[name] = df[col].max().strftime("%Y-%m-%d")
                    break
    return max_mod_date


# def split_df_on_nan_rows(df):
#     nan_mask = df.isnull().all(axis=1)
#     groups = df[~nan_mask].groupby(
#         (nan_mask.shift(1, fill_value=False) | nan_mask).cumsum()
#     )
#     split_dfs = [group for _, group in groups]
#     return split_dfs


# Function to set the first row as header for all but the first DataFrame
# def set_first_row_as_header(dfs):
#     for i in range(1, len(dfs)):
#         # dfs[i].columns = dfs[i].iloc[0]
#         dfs[i] = dfs[i][1:]
#     return dfs


# def get_new_start_cell(start_cell, df):
#     start_row, start_col = coordinate_to_tuple(start_cell)
#     new_start_row = start_row + len(df) + 1
#     return f"{get_column_letter(start_col)}{new_start_row}"


# def paste_data(df, start_cell, ws, header=True):
#     # Extract row and column indices from the start_cell
#     start_row, start_col = coordinate_to_tuple(start_cell)

#     # Write DataFrame to worksheet starting from the start_cell
#     for r_idx, row in enumerate(
#         dataframe_to_rows(df, index=False, header=header), start_row
#     ):
#         for c_idx, value in enumerate(row, start_col):
#             cell = ws[f"{get_column_letter(c_idx)}{r_idx}"]
#             cell.value = value
#             # print(f"Pasting value {value} to cell {get_column_letter(c_idx)}{r_idx}")
#             if r_idx == start_row:  # Check if it's the header row
#                 cell.font = Font(bold=True)  # Set font to bold
#     return get_new_start_cell(start_cell, df)
